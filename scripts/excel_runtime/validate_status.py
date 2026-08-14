import sys
import openpyxl
from openpyxl.utils import range_boundaries

TARGET_TABLES = {
    "tblContas",
    "tblCartoes",
    "tblLancamentos",
    "tblOrcamento",
    "tblMetas",
    "tblReserva",
    "tblInvestimentos",
    "tblBensPatrimoniais",
    "tblDividas"
}

def validate_workbook_status(filepath):
    print(f"Validating semantic Status for {filepath}...")
    wb_formulas = openpyxl.load_workbook(filepath, data_only=False)
    wb_values = openpyxl.load_workbook(filepath, data_only=True)
    
    errors = []
    
    for sheet_name in wb_formulas.sheetnames:
        ws_f = wb_formulas[sheet_name]
        ws_v = wb_values[sheet_name]
        
        for tbl in ws_f.tables.values():
            if tbl.name not in TARGET_TABLES:
                continue
                
            # Find Status column index
            status_col_idx = None
            for idx, col in enumerate(tbl.tableColumns):
                if col.name == "Status":
                    status_col_idx = idx
                    break
            
            if status_col_idx is None:
                continue
                
            min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
            header_rows = tbl.headerRowCount if tbl.headerRowCount is not None else 1
            totals_rows = tbl.totalsRowCount if tbl.totalsRowCount is not None else 0
            
            start_row = min_row + header_rows
            end_row = max_row - totals_rows
            
            for r in range(start_row, end_row + 1):
                # Check if row is populated using ws_f
                is_populated = False
                for c in range(min_col, max_col + 1):
                    tc = ws_f.cell(row=r, column=c)
                    if tc.data_type != 'f' and not (isinstance(tc.value, str) and tc.value.startswith('=')):
                        if tc.value not in (None, ""):
                            is_populated = True
                            break
                        
                if is_populated:
                    status_val = ws_v.cell(row=r, column=min_col + status_col_idx).value
                    if status_val != "OK":
                        errors.append({
                            "sheet": sheet_name,
                            "table": tbl.name,
                            "row": r,
                            "status": status_val
                        })
                        
    if errors:
        print(f"\nFOUND {len(errors)} STATUS ERRORS in {filepath}:")
        for err in errors:
            print(f"Sheet: {err['sheet']} | Table: {err['table']} | Row: {err['row']} | Status: {err['status']}")
        return False
        
    print(f"Status validation passed for {filepath}.")
    return True

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: validate_status.py <workbook.xlsx>")
        sys.exit(1)

    filepath = sys.argv[1]
    success = validate_workbook_status(filepath)
    if not success:
        sys.exit(1)
    sys.exit(0)
