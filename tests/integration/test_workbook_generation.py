import os
import openpyxl
from excel_saas.application.generate_workbook import generate
from excel_saas.core.models.generation_request import GenerationRequest

def test_full_pipeline_light_and_dark(tmp_path):
    output_dir = str(tmp_path)

    req_light = GenerationRequest(template_id="finance_personal", year=2026, theme="light")
    path_light = generate(req_light, output_dir)

    assert os.path.exists(path_light)

    # Verify via openpyxl
    wb = openpyxl.load_workbook(path_light, data_only=False)
    sheet_names = wb.sheetnames

    assert sheet_names == [
        "Comece Aqui",
        "Dashboard",
        "Lançamentos",
        "Contas",
        "Cartões",
        "Faturas",
        "Parcelamentos",
        "Orçamento",
        "Metas",
        "Reserva",
        "Investimentos",
        "Dashboard Investimentos",
        "Patrimônio",
        "Dashboard Patrimônio",
        "Dívidas",
        "Configurações"
    ]

    ws = wb["Lançamentos"]

    # Check table existence (openpyxl supports this)
    assert "tblLancamentos" in ws.tables

    table = ws.tables["tblLancamentos"]
    assert table.ref == "B4:AF5" # 31 columns now

    lancamentos_headers = [cell.value for cell in ws[4] if cell.value]
    assert lancamentos_headers == [
        "Data", "Descrição", "Tipo", "Categoria", "Conta", "Conta destino",
        "Cartão", "Competência da fatura", "Valor", "Parcelas",
        "Essencial?", "Recorrente?", "Status", "Status fatura", "Observação",
        "sys_CompetenciaEfetiva", "sys_Receita", "sys_Despesa",
        "sys_CaixaConta", "sys_CaixaDestino", "sys_Cartao",
        "sys_ValorParcela", "sys_CompromissoFuturo",
        "sys_ParcelasEfetivas", "sys_FaturaInicial", "sys_FaturaFinal",
        "sys_ValorParcelaBase", "sys_ValorUltimaParcela", "sys_AjusteUltimaParcela",
        "sys_CreditoFatura", "sys_PagamentoFatura"
    ]

    comp_fatura_cell = ws["I5"]
    assert "mmm/yyyy" in comp_fatura_cell.number_format or comp_fatura_cell.number_format == "General"
    # Verify Phase 5B formula cells and data_type
    phase5b_formula_cells = {
        "Y5": "sys_ParcelasEfetivas",
        "Z5": "sys_FaturaInicial",
        "AA5": "sys_FaturaFinal",
        "AB5": "sys_ValorParcelaBase",
        "AC5": "sys_ValorUltimaParcela",
        "AD5": "sys_AjusteUltimaParcela",
        "AE5": "sys_CreditoFatura",
        "AF5": "sys_PagamentoFatura"
    }

    for cell_ref, col_name in phase5b_formula_cells.items():
        cell = ws[cell_ref]
        assert str(cell.value).startswith("="), f"{col_name} at {cell_ref} must be a formula"
        assert cell.data_type == "f", f"{col_name} at {cell_ref} must have data_type 'f'"

    # Verify physical hidden state for columns Y through AF (indices 25 to 32, assuming A=1)
    from openpyxl.utils import column_index_from_string
    hidden_indices = {column_index_from_string(col) for col in ["Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF"]}

    covered_hidden = set()
    for col_dim in ws.column_dimensions.values():
        if getattr(col_dim, "hidden", False):
            # min and max are column indices
            dim_min = getattr(col_dim, "min", None)
            dim_max = getattr(col_dim, "max", None)
            if dim_min and dim_max:
                for idx in range(dim_min, dim_max + 1):
                    covered_hidden.add(idx)

    # Assert all target Phase 5B columns are covered by hidden dimensions
    for col_idx in hidden_indices:
        assert col_idx in covered_hidden, f"Column index {col_idx} is not hidden in the generated xlsx."


    # Check new tables
    contas_ws = wb["Contas"]
    assert "tblContas" in contas_ws.tables

    cartoes_ws = wb["Cartões"]
    assert "tblCartoes" in cartoes_ws.tables

    # Check exact columns for tblContas and tblCartoes
    contas_headers = [cell.value for cell in contas_ws[4] if cell.value]
    assert contas_headers == ["Nome", "Tipo", "Instituição", "Saldo inicial", "Saldo atual", "Incluir no saldo disponível?", "Ativa?", "Status"]

    cartoes_headers = [cell.value for cell in cartoes_ws[6] if cell.value]
    assert cartoes_headers == ["Nome", "Limite", "Dia fechamento", "Dia vencimento", "Conta de pagamento", "Ativo?", "Status", "sys_DiaFechamentoSeguro", "sys_DiaVencimentoSeguro"]


    # Phase 5C: Check Faturas table
    faturas_ws = wb["Faturas"]
    assert "tblFaturas" in faturas_ws.tables
    
    table_f = faturas_ws.tables["tblFaturas"]
    assert table_f.ref == "B4:M5"

    faturas_headers = [cell.value for cell in faturas_ws[4] if cell.value]
    assert faturas_headers == [
        "Cartão", "Competência", "Fechamento", "Vencimento", "Compras / Parcelas",
        "Créditos / Estornos", "Total da fatura", "Pagamentos", "Em aberto",
        "Status", "Situação", "sys_CompetenciaNormalizada"
    ]
    
    # Check blank data row formula cells
    for col in ["D5", "E5", "F5", "G5", "H5", "I5", "J5", "K5", "L5", "M5"]:
        cell = faturas_ws[col]
        assert str(cell.value).startswith("="), f"Cell {col} should be formula, got {cell.value}"
        assert cell.data_type == "f", f"Cell {col} should have data_type 'f', got {cell.data_type}"

    # Check M column hidden physically
    from openpyxl.utils import column_index_from_string
    col_m_idx = column_index_from_string("M")
    m_hidden = False
    for col_dim in faturas_ws.column_dimensions.values():
        if getattr(col_dim, "hidden", False):
            d_min = getattr(col_dim, "min", None)
            d_max = getattr(col_dim, "max", None)
            if d_min and d_max and d_min <= col_m_idx <= d_max:
                m_hidden = True
                break
    assert m_hidden, "Column M in Faturas is not physically hidden"

    # Check validations
    cartao_val = None
    comp_val = None
    for val in faturas_ws.data_validations.dataValidation:
        sqref = val.sqref.__str__()
        if "B" in sqref:
            cartao_val = val
        elif "C" in sqref:
            comp_val = val

    assert cartao_val is not None
    assert cartao_val.type == "list"
    assert "lista_cartoes" in cartao_val.formula1

    assert comp_val is not None
    assert comp_val.type == "date"
    assert comp_val.operator in ("between", None)
    assert "1" in comp_val.formula1
    assert "2958465" in comp_val.formula2
    assert getattr(comp_val, "allowBlank", False)

    # Check number formats
    assert "mmm/yyyy" in faturas_ws["C5"].number_format or faturas_ws["C5"].number_format == "General"
    assert "dd/mm/yyyy" in faturas_ws["D5"].number_format
    assert "dd/mm/yyyy" in faturas_ws["E5"].number_format
    for col in ["F", "G", "H", "I", "J"]:
        fmt = faturas_ws[f"{col}5"].number_format
        assert "R$" in fmt or "General" not in fmt # Just ensuring it's formatting

    # Phase 5D: Check Parcelamentos table
    parcelamentos_ws = wb["Parcelamentos"]
    assert "tblParcelamentos" in parcelamentos_ws.tables
    
    table_p = parcelamentos_ws.tables["tblParcelamentos"]
    assert table_p.ref == "B4:O104"
    
    parcelamentos_headers = [cell.value for cell in parcelamentos_ws[4] if cell.value]
    assert parcelamentos_headers == [
        "Descrição", "Cartão", "Valor original", "Parcelas",
        "Valor da parcela", "Primeira fatura", "Última fatura",
        "Parcelas restantes", "Próxima competência",
        "Compromisso restante", "Situação",
        "sys_Ordem", "sys_IndiceLancamento", "sys_AjusteUltimaParcela"
    ]
    
    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "N", "O"]:
        cell = parcelamentos_ws[f"{col}5"]
        assert str(cell.value).startswith("=")
        assert cell.data_type == "f"
    
    assert parcelamentos_ws["M5"].value == 1
    
    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "N", "O"]:
        cell = parcelamentos_ws[f"{col}104"]
        assert str(cell.value).startswith("=")
        assert cell.data_type == "f"
        
    assert parcelamentos_ws["M104"].value == 100
    
    target_hidden_cols = {column_index_from_string("M"), column_index_from_string("N"), column_index_from_string("O")}
    covered_hidden_p = set()
    for col_dim in parcelamentos_ws.column_dimensions.values():
        if getattr(col_dim, "hidden", False):
            d_min = getattr(col_dim, "min", None)
            d_max = getattr(col_dim, "max", None)
            if d_min and d_max:
                for idx in range(d_min, d_max + 1):
                    covered_hidden_p.add(idx)
    
    for col_idx in target_hidden_cols:
        assert col_idx in covered_hidden_p
        
    assert "R$" in parcelamentos_ws["D5"].number_format
    assert "R$" in parcelamentos_ws["F5"].number_format
    assert "mmm/yyyy" in parcelamentos_ws["G5"].number_format or parcelamentos_ws["G5"].number_format == "General"
    assert "mmm/yyyy" in parcelamentos_ws["H5"].number_format or parcelamentos_ws["H5"].number_format == "General"
    assert "mmm/yyyy" in parcelamentos_ws["J5"].number_format or parcelamentos_ws["J5"].number_format == "General"
    assert "R$" in parcelamentos_ws["K5"].number_format
    
    assert parcelamentos_ws.protection.sheet is True

    # Phase 6: Check Orçamento table
    orcamento_ws = wb["Orçamento"]
    assert "tblOrcamento" in orcamento_ws.tables
    
    table_orc = orcamento_ws.tables["tblOrcamento"]
    assert table_orc.ref == "B4:L5"
    
    orcamento_headers = [cell.value for cell in orcamento_ws[4] if cell.value]
    assert orcamento_headers == [
        "Competência", "Categoria", "Orçamento", "Saídas em conta", "Cartão / Parcelas",
        "Consumido / Comprometido", "Disponível", "Uso %", "Status", "Situação", "sys_CompetenciaNormalizada"
    ]
    
    for col in ["E", "F", "G", "H", "I", "J", "K", "L"]:
        cell = orcamento_ws[f"{col}5"]
        assert str(cell.value).startswith("=")
        assert cell.data_type == "f"
    
    # Validations
    comp_val = None
    cat_val = None
    orc_val = None
    for val in orcamento_ws.data_validations.dataValidation:
        sqref = val.sqref.__str__()
        if "B" in sqref:
            comp_val = val
        elif "C" in sqref:
            cat_val = val
        elif "D" in sqref:
            orc_val = val

    assert comp_val is not None
    assert comp_val.type == "date"
    assert comp_val.operator in (None, "between")
    assert "1" in comp_val.formula1
    assert "2958465" in comp_val.formula2
    assert comp_val.allowBlank is True
    
    assert cat_val is not None
    assert cat_val.type == "list"
    assert "lista_categorias" in cat_val.formula1
    assert cat_val.allowBlank is True
    
    assert orc_val is not None
    assert orc_val.type == "decimal"
    assert orc_val.operator == "greaterThanOrEqual"
    assert "0" in orc_val.formula1
    assert orc_val.allowBlank is True
    
    # Hidden system column L
    from openpyxl.utils import column_index_from_string
    col_l_idx = column_index_from_string("L")
    covered_hidden_orc = False
    for col_dim in orcamento_ws.column_dimensions.values():
        if getattr(col_dim, "hidden", False):
            d_min = getattr(col_dim, "min", None)
            d_max = getattr(col_dim, "max", None)
            if d_min and d_max and d_min <= col_l_idx <= d_max:
                covered_hidden_orc = True
                break
    assert covered_hidden_orc, "Column L in Orçamento is not physically hidden"
    
    # Formats
    assert "mmm/yyyy" in orcamento_ws["B5"].number_format or orcamento_ws["B5"].number_format == "General"
    assert "R$" in orcamento_ws["D5"].number_format or orcamento_ws["D5"].number_format == "General"
    assert "R$" in orcamento_ws["E5"].number_format or orcamento_ws["E5"].number_format == "General"
    assert "R$" in orcamento_ws["F5"].number_format or orcamento_ws["F5"].number_format == "General"
    assert "R$" in orcamento_ws["G5"].number_format or orcamento_ws["G5"].number_format == "General"
    assert "R$" in orcamento_ws["H5"].number_format or orcamento_ws["H5"].number_format == "General"
    assert "0.0%" in orcamento_ws["I5"].number_format or orcamento_ws["I5"].number_format == "General"
    
    # Protection
    assert orcamento_ws.protection.sheet is False

    # Phase 7: Check Metas table
    metas_ws = wb["Metas"]
    assert "tblMetas" in metas_ws.tables
    
    table_metas = metas_ws.tables["tblMetas"]
    assert table_metas.ref == "B4:K5"
    
    metas_headers = [cell.value for cell in metas_ws[4] if cell.value]
    assert metas_headers == [
        "Meta", "Valor alvo", "Valor atual", "Data alvo", "Falta",
        "Progresso %", "Meses restantes", "Aporte mensal necessário", "Status", "Situação"
    ]
    
    for col in ["F", "G", "H", "I", "J", "K"]:
        cell = metas_ws[f"{col}5"]
        assert str(cell.value).startswith("=")
        assert cell.data_type == "f"
        
    for col in ["B", "C", "D", "E"]:
        cell = metas_ws[f"{col}5"]
        assert not str(cell.value).startswith("=")
        
    alvo_val = None
    atual_val = None
    data_alvo_val = None
    for val in metas_ws.data_validations.dataValidation:
        sqref = val.sqref.__str__()
        if "C" in sqref:
            alvo_val = val
        elif "D" in sqref:
            atual_val = val
        elif "E" in sqref:
            data_alvo_val = val
            
    assert alvo_val is not None
    assert alvo_val.type == "decimal"
    assert alvo_val.operator == "greaterThan"
    assert "0" in alvo_val.formula1
    
    assert atual_val is not None
    assert atual_val.type == "decimal"
    assert atual_val.operator == "greaterThanOrEqual"
    assert "0" in atual_val.formula1
    
    assert data_alvo_val is not None
    assert data_alvo_val.type == "date"
    assert data_alvo_val.operator in (None, "between")
    assert "1" in data_alvo_val.formula1
    assert "2958465" in data_alvo_val.formula2
    
    assert "R$" in metas_ws["C5"].number_format or metas_ws["C5"].number_format == "General"
    assert "R$" in metas_ws["D5"].number_format or metas_ws["D5"].number_format == "General"
    assert "mmm/yyyy" in metas_ws["E5"].number_format or metas_ws["E5"].number_format == "General"
    assert "R$" in metas_ws["F5"].number_format or metas_ws["F5"].number_format == "General"
    assert "0.0%" in metas_ws["G5"].number_format or metas_ws["G5"].number_format == "General"
    assert "R$" in metas_ws["I5"].number_format or metas_ws["I5"].number_format == "General"
    
    assert metas_ws.protection.sheet is False

    # Phase 8: Check Reserva table
    reserva_ws = wb["Reserva"]
    assert "tblReserva" in reserva_ws.tables
    
    table_reserva = reserva_ws.tables["tblReserva"]
    assert table_reserva.ref == "B4:J5"
    
    reserva_headers = [cell.value for cell in reserva_ws[4] if cell.value]
    assert reserva_headers == [
        "Custo essencial mensal", "Meses desejados", "Reserva atual", "Reserva alvo",
        "Falta", "Cobertura atual", "Progresso %", "Status", "Situação"
    ]
    
    for col in ["E", "F", "G", "H", "I", "J"]:
        cell = reserva_ws[f"{col}5"]
        assert str(cell.value).startswith("=")
        assert cell.data_type == "f"
        
    for col in ["B", "C", "D"]:
        cell = reserva_ws[f"{col}5"]
        assert not str(cell.value).startswith("=")
        
    custo_val = None
    meses_val = None
    reserva_val = None
    for val in reserva_ws.data_validations.dataValidation:
        sqref = val.sqref.__str__()
        if "B" in sqref:
            custo_val = val
        elif "C" in sqref:
            meses_val = val
        elif "D" in sqref:
            reserva_val = val
            
    assert custo_val is not None
    assert custo_val.type == "decimal"
    assert custo_val.operator == "greaterThan"
    assert "0" in custo_val.formula1
    
    assert meses_val is not None
    assert meses_val.type == "whole"
    assert meses_val.operator == "greaterThanOrEqual"
    assert "1" in meses_val.formula1
    
    assert reserva_val is not None
    assert reserva_val.type == "decimal"
    assert reserva_val.operator == "greaterThanOrEqual"
    assert "0" in reserva_val.formula1
    
    assert "R$" in reserva_ws["B5"].number_format or reserva_ws["B5"].number_format == "General"
    assert "General" in reserva_ws["C5"].number_format or reserva_ws["C5"].number_format == "0"
    assert "R$" in reserva_ws["D5"].number_format or reserva_ws["D5"].number_format == "General"
    assert "R$" in reserva_ws["E5"].number_format or reserva_ws["E5"].number_format == "General"
    assert "R$" in reserva_ws["F5"].number_format or reserva_ws["F5"].number_format == "General"
    assert "0.0" in reserva_ws["G5"].number_format or reserva_ws["G5"].number_format == "General"
    assert "0.0%" in reserva_ws["H5"].number_format or reserva_ws["H5"].number_format == "General"
    
    assert reserva_ws.protection.sheet is False

    # Check Seed default
    assert reserva_ws["B5"].value is None
    assert reserva_ws["C5"].value == 6
    assert reserva_ws["D5"].value is None

    # Phase 9: Check Investimentos table
    investimentos_ws = wb["Investimentos"]
    assert "tblInvestimentos" in investimentos_ws.tables
    
    table_investimentos = investimentos_ws.tables["tblInvestimentos"]
    assert table_investimentos.ref == "B4:O5"
    
    invest_headers = [cell.value for cell in investimentos_ws[4] if cell.value]
    assert invest_headers == [
        "Ativo", "Classe", "Instituição", "Total aportado", "Total recebido", "Valor atual",
        "Resultado total", "Retorno simples %", "Peso carteira %", "Status", "Situação",
        "sys_AporteValido", "sys_RecebidoValido", "sys_ValorAtualValido"
    ]
    
    for col in ["H", "I", "J", "K", "L", "M", "N", "O"]:
        cell = investimentos_ws[f"{col}5"]
        assert str(cell.value).startswith("=")
        assert cell.data_type == "f"
        
    for col in ["B", "C", "D", "E", "F", "G"]:
        cell = investimentos_ws[f"{col}5"]
        assert not str(cell.value).startswith("=")
        
    classe_val = None
    aportado_val = None
    recebido_val = None
    atual_val = None
    for val in investimentos_ws.data_validations.dataValidation:
        sqref = val.sqref.__str__()
        if "C" in sqref:
            classe_val = val
        elif "E" in sqref:
            aportado_val = val
        elif "F" in sqref:
            recebido_val = val
        elif "G" in sqref:
            atual_val = val
            
    assert classe_val is not None
    assert classe_val.type == "list"
    assert '"Renda fixa' in classe_val.formula1
    
    assert aportado_val is not None
    assert aportado_val.type == "decimal"
    assert aportado_val.operator == "greaterThan"
    assert "0" in aportado_val.formula1
    
    assert recebido_val is not None
    assert recebido_val.type == "decimal"
    assert recebido_val.operator == "greaterThanOrEqual"
    assert "0" in recebido_val.formula1
    
    assert atual_val is not None
    assert atual_val.type == "decimal"
    assert atual_val.operator == "greaterThanOrEqual"
    assert "0" in atual_val.formula1
    
    assert "R$" in investimentos_ws["E5"].number_format or investimentos_ws["E5"].number_format == "General"
    assert "R$" in investimentos_ws["F5"].number_format or investimentos_ws["F5"].number_format == "General"
    assert "R$" in investimentos_ws["G5"].number_format or investimentos_ws["G5"].number_format == "General"
    assert "R$" in investimentos_ws["H5"].number_format or investimentos_ws["H5"].number_format == "General"
    assert "0.0%" in investimentos_ws["I5"].number_format or investimentos_ws["I5"].number_format == "General"
    assert "0.0%" in investimentos_ws["J5"].number_format or investimentos_ws["J5"].number_format == "General"
    
    assert investimentos_ws.protection.sheet is False

    # Phase 10: Check Dashboard Investimentos
    dash_inv_ws = wb["Dashboard Investimentos"]
    assert dash_inv_ws.protection.sheet is True
    
    assert dash_inv_ws["B4"].value == "Total aportado"
    assert dash_inv_ws["D4"].value == "Total recebido"
    assert dash_inv_ws["F4"].value == "Valor atual"
    assert dash_inv_ws["B7"].value == "Resultado total"
    assert dash_inv_ws["D7"].value == "Retorno simples %"
    
    for cell in ["B5", "D5", "F5", "B8", "D8"]:
        assert str(dash_inv_ws[cell].value).startswith("=")
        assert dash_inv_ws[cell].data_type == "f"
        
    for cell in ["B5", "D5", "F5", "B8"]:
        assert "R$" in dash_inv_ws[cell].number_format or dash_inv_ws[cell].number_format == "General"
    assert "0.0%" in dash_inv_ws["D8"].number_format or dash_inv_ws["D8"].number_format == "General"
    
    assert "tblResumoInvestimentos" in dash_inv_ws.tables
    table_resumo = dash_inv_ws.tables["tblResumoInvestimentos"]
    assert table_resumo.ref == "B11:H19"
    
    resumo_headers = [cell.value for cell in dash_inv_ws[11] if cell.value]
    assert resumo_headers == [
        "Classe", "Total aportado", "Total recebido", "Valor atual", 
        "Resultado total", "Peso carteira %", "Retorno simples %"
    ]
    
    classes_esperadas = ["Renda fixa", "Ações", "FIIs", "ETFs", "Fundos", "Cripto", "Previdência", "Outros"]
    for i, expected_class in enumerate(classes_esperadas, start=12):
        cell = dash_inv_ws[f"B{i}"]
        assert cell.value == expected_class
        assert cell.data_type != "f"
        assert cell.protection.locked is True
        
    for i in range(12, 20):
        for col in ["C", "D", "E", "F", "G", "H"]:
            cell = dash_inv_ws[f"{col}{i}"]
            assert str(cell.value).startswith("=")
            assert cell.data_type == "f"

    from openpyxl.utils import column_index_from_string
    target_hidden_cols = {column_index_from_string("M"), column_index_from_string("N"), column_index_from_string("O")}
    covered_hidden_inv = set()
    for _, col_dim in investimentos_ws.column_dimensions.items():
        if getattr(col_dim, "hidden", False):
            for idx in range(col_dim.min, col_dim.max + 1):
                covered_hidden_inv.add(idx)
                
    for col_idx in target_hidden_cols:
        assert col_idx in covered_hidden_inv

    # Phase 11: Check Patrimônio
    patrimonio_ws = wb["Patrimônio"]
    assert patrimonio_ws.protection.sheet is False

    assert "tblBensPatrimoniais" in patrimonio_ws.tables
    table_bens = patrimonio_ws.tables["tblBensPatrimoniais"]
    assert table_bens.ref == "B4:G5"

    bens_headers = [cell.value for cell in patrimonio_ws[4] if cell.value]
    assert bens_headers == ["Bem", "Categoria", "Valor atual", "Observação", "Status", "sys_ValorAtualValido"]

    for col in ["B", "C", "D", "E"]:
        cell = patrimonio_ws[f"{col}5"]
        assert not str(cell.value).startswith("=")
        
    for col in ["F", "G"]:
        cell = patrimonio_ws[f"{col}5"]
        assert str(cell.value).startswith("=")
        assert cell.data_type == "f"

    cat_val_bens = None
    valor_val_bens = None
    for val in patrimonio_ws.data_validations.dataValidation:
        sqref = val.sqref.__str__()
        if "C" in sqref:
            cat_val_bens = val
        elif "D" in sqref:
            valor_val_bens = val

    assert cat_val_bens is not None
    assert cat_val_bens.type == "list"
    assert '"Imóveis' in cat_val_bens.formula1
    assert 'Outros"' in cat_val_bens.formula1

    assert valor_val_bens is not None
    assert valor_val_bens.type == "decimal"
    assert valor_val_bens.operator == "greaterThanOrEqual"
    assert "0" in valor_val_bens.formula1

    assert "R$" in patrimonio_ws["D5"].number_format or patrimonio_ws["D5"].number_format == "General"

    target_hidden_cols_patrimonio = {column_index_from_string("G")}
    covered_hidden_pat = set()
    for _, col_dim in patrimonio_ws.column_dimensions.items():
        if getattr(col_dim, "hidden", False):
            for idx in range(col_dim.min, col_dim.max + 1):
                covered_hidden_pat.add(idx)
                
    for col_idx in target_hidden_cols_patrimonio:
        assert col_idx in covered_hidden_pat

    # Phase 12: Check Dashboard Patrimônio
    dash_pat_ws = wb["Dashboard Patrimônio"]
    assert dash_pat_ws.protection.sheet is True

    assert "tblResumoPatrimonio" in dash_pat_ws.tables
    table_dash_pat = dash_pat_ws.tables["tblResumoPatrimonio"]
    assert table_dash_pat.ref == "B11:D14"

    dash_pat_headers = [cell.value for cell in dash_pat_ws[11] if cell.value]
    assert dash_pat_headers == ["Componente", "Valor atual", "Peso %"]

    assert dash_pat_ws["B2"].value == "Dashboard Patrimônio"
    assert dash_pat_ws["B3"].value == "Visão consolidada dos ativos atuais. Dívidas e outros passivos ainda não são descontados."
    assert dash_pat_ws["B4"].value == "Contas e caixa"
    assert dash_pat_ws["D4"].value == "Investimentos"
    assert dash_pat_ws["F4"].value == "Bens patrimoniais"
    assert dash_pat_ws["B7"].value == "Ativos totais"
    assert dash_pat_ws["B9"].value == "Contas representam caixa; investimentos e bens devem ser cadastrados apenas em suas abas próprias para evitar dupla contagem."

    for col_row in ["B5", "D5", "F5", "B8"]:
        cell = dash_pat_ws[col_row]
        assert str(cell.value).startswith("=")
        assert cell.data_type == "f"
        assert cell.number_format == "R$ #,##0.00"
        
    for row, expected_label in zip(range(12, 15), ["Contas e caixa", "Investimentos", "Bens patrimoniais"]):
        cell_b = dash_pat_ws[f"B{row}"]
        assert cell_b.value == expected_label
        assert cell_b.data_type != "f"
        assert cell_b.protection.locked is True
        
        assert str(dash_pat_ws[f"C{row}"].value).startswith("=")
        assert dash_pat_ws[f"C{row}"].data_type == "f"
        assert dash_pat_ws[f"C{row}"].number_format == "R$ #,##0.00"
        
        assert str(dash_pat_ws[f"D{row}"].value).startswith("=")
        assert dash_pat_ws[f"D{row}"].data_type == "f"
        assert dash_pat_ws[f"D{row}"].number_format == "0.0%"

    config_ws = wb["Configurações"]
    assert "tblCategorias" in config_ws.tables

    # Phase 13: Check Dívidas
    dividas_ws = wb["Dívidas"]
    assert dividas_ws.protection.sheet is False
    assert dividas_ws.freeze_panes == "B5"

    assert dividas_ws["B2"].value == "Dívidas"
    assert dividas_ws["B3"].value == "Cadastre empréstimos, financiamentos e outras dívidas contratuais. Não repita faturas, compras parceladas ou saldos negativos já registrados em outras abas."

    assert "tblDividas" in dividas_ws.tables
    table_dividas = dividas_ws.tables["tblDividas"]
    assert table_dividas.ref == "B4:K5"
    
    dividas_headers = [cell.value for cell in dividas_ws[4] if cell.value]
    assert dividas_headers == [
        "Dívida", "Categoria", "Credor", "Saldo devedor atual",
        "Parcela mensal atual", "Data final", "Observação",
        "Status", "sys_SaldoDevedorValido", "sys_ParcelaMensalValida"
    ]

    for col in ["B", "C", "D", "E", "F", "G", "H"]:
        cell = dividas_ws[f"{col}5"]
        assert not str(cell.value).startswith("=")

    for col in ["I", "J", "K"]:
        cell = dividas_ws[f"{col}5"]
        assert str(cell.value).startswith("=")
        assert cell.data_type == "f"

    cat_val_dividas = None
    saldo_val_dividas = None
    parcela_val_dividas = None
    data_val_dividas = None
    
    for val in dividas_ws.data_validations.dataValidation:
        sqref = val.sqref.__str__()
        if "C" in sqref:
            cat_val_dividas = val
        elif "E" in sqref:
            saldo_val_dividas = val
        elif "F" in sqref:
            parcela_val_dividas = val
        elif "G" in sqref:
            data_val_dividas = val
            
    assert cat_val_dividas is not None
    assert cat_val_dividas.type == "list"
    assert cat_val_dividas.formula1 == '"Financiamento imobiliário,Financiamento de veículo,Empréstimo pessoal,Consignado,Dívida tributária,Dívida com pessoa,Outros"'

    assert saldo_val_dividas is not None
    assert saldo_val_dividas.type == "decimal"
    assert saldo_val_dividas.operator == "greaterThan"
    assert saldo_val_dividas.formula1 == "0"
    
    assert parcela_val_dividas is not None
    assert parcela_val_dividas.type == "decimal"
    assert parcela_val_dividas.operator == "greaterThanOrEqual"
    assert parcela_val_dividas.formula1 == "0"
    
    assert data_val_dividas is not None
    assert data_val_dividas.type == "date"
    if data_val_dividas.operator is None:
        data_val_dividas.operator = "between"
    assert data_val_dividas.operator == "between"
    assert data_val_dividas.formula1 == "1"
    assert data_val_dividas.formula2 == "2958465"

    assert dividas_ws["E5"].number_format == "R$ #,##0.00"
    assert dividas_ws["F5"].number_format == "R$ #,##0.00"
    assert dividas_ws["J5"].number_format == "R$ #,##0.00"
    assert dividas_ws["K5"].number_format == "R$ #,##0.00"

    target_hidden_cols_dividas = {column_index_from_string("J"), column_index_from_string("K")}
    covered_hidden_div = set()
    for _, col_dim in dividas_ws.column_dimensions.items():
        if getattr(col_dim, "hidden", False):
            for idx in range(col_dim.min, col_dim.max + 1):
                covered_hidden_div.add(idx)

    for col_idx in target_hidden_cols_dividas:
        assert col_idx in covered_hidden_div
    assert "tblTiposConta" in config_ws.tables
    assert "tblContas" not in config_ws.tables
    assert "tblCartoes" not in config_ws.tables

    # Check that protection was disabled for natural expansion
    assert ws.protection.sheet is False
    assert contas_ws.protection.sheet is False
    assert cartoes_ws.protection.sheet is False
    assert faturas_ws.protection.sheet is False
    assert config_ws.protection.sheet is False

    assert wb["Comece Aqui"].protection.sheet is True
    assert wb["Dashboard"].protection.sheet is True

    # Verify defined names
    assert "lista_categorias" in wb.defined_names
    assert "lista_contas" in wb.defined_names
    assert "lista_cartoes" in wb.defined_names
    assert "lista_tipos_conta" in wb.defined_names

    assert wb.defined_names["lista_categorias"].value == "tblCategorias[Categoria]"
    assert wb.defined_names["lista_tipos_conta"].value == "tblTiposConta[Tipo]"
    assert wb.defined_names["lista_contas"].value == "tblContas[Nome]"
    assert wb.defined_names["lista_cartoes"].value == "tblCartoes[Nome]"

    # Bug 1: Verify Dashboard formulas are written as formulas, not text
    dash_ws = wb["Dashboard"]

    # Saldo Disponível Hoje
    saldo_disponivel = dash_ws["B5"]
    assert str(saldo_disponivel.value).startswith("=")
    assert "SUMIFS(tblContas[Saldo atual]" in str(saldo_disponivel.value)

    # Resultado Registrado
    saldo_cell = dash_ws["B8"]
    assert str(saldo_cell.value).startswith("=")
    assert "SUM(tblLancamentos[sys_Receita])" in str(saldo_cell.value)

    # Receitas Registradas
    receita_cell = dash_ws["D8"]
    assert str(receita_cell.value).startswith("=")
    assert "SUM(tblLancamentos[sys_Receita])" in str(receita_cell.value)

    # Despesas Registradas
    despesa_cell = dash_ws["E8"]
    assert str(despesa_cell.value).startswith("=")
    assert "SUM(tblLancamentos[sys_Despesa])" in str(despesa_cell.value)

    # Check data validation logic for Contas (especially Saldo inicial on column E)
    saldo_val = None
    for val in contas_ws.data_validations.dataValidation:
        if "E" in val.sqref.__str__():
            saldo_val = val
            break

    assert saldo_val is not None
    assert saldo_val.type == "decimal"
    assert saldo_val.operator in (None, "between") # between is the default operator in Excel XML
    assert saldo_val.allowBlank is True

    # We must assert that it correctly handles negative, positive, and zero, and uses exact Excel technical bounds
    assert "-9.99999999999999e+307" in saldo_val.formula1.lower()
    assert "9.99999999999999e+307" in saldo_val.formula2.lower()
    assert "1000000000" not in saldo_val.formula1

    # Check Limite >= 0
    limite_val = None
    dia_fechamento_val = None
    dia_vencimento_val = None
    conta_pagamento_val = None
    for val in cartoes_ws.data_validations.dataValidation:
        sqref = val.sqref.__str__()
        if "C" in sqref:
            limite_val = val
        elif "D" in sqref:
            dia_fechamento_val = val
        elif "E" in sqref:
            dia_vencimento_val = val
        elif "F" in sqref:
            conta_pagamento_val = val

    assert limite_val is not None
    assert limite_val.type == "decimal"
    assert "0" in limite_val.formula1

    assert dia_fechamento_val is not None
    assert dia_fechamento_val.type == "whole"
    assert "1" in dia_fechamento_val.formula1
    assert "31" in dia_fechamento_val.formula2

    assert dia_vencimento_val is not None
    assert dia_vencimento_val.type == "whole"
    assert "1" in dia_vencimento_val.formula1
    assert "31" in dia_vencimento_val.formula2

    assert conta_pagamento_val is not None
    assert conta_pagamento_val.type == "list"
    assert "lista_contas" in conta_pagamento_val.formula1

    # Check Lançamentos validations
    categoria_val = None
    conta_val = None
    conta_dest_val = None
    cartao_val = None
    valor_val = None
    parcelas_val = None

    for val in ws.data_validations.dataValidation:
        sqref = val.sqref.__str__()
        if "E" in sqref:
            categoria_val = val
        elif "F" in sqref:
            conta_val = val
        elif "G" in sqref:
            conta_dest_val = val
        elif "H" in sqref:
            cartao_val = val
        elif "J" in sqref:
            valor_val = val
        elif "K" in sqref:
            parcelas_val = val

    assert categoria_val is not None and "lista_categorias" in categoria_val.formula1
    assert conta_val is not None and "lista_contas" in conta_val.formula1
    assert conta_dest_val is not None and "lista_contas" in conta_dest_val.formula1
    assert cartao_val is not None and "lista_cartoes" in cartao_val.formula1

    assert valor_val is not None
    assert valor_val.type == "decimal"
    assert valor_val.operator == "greaterThan"
    assert "0" in valor_val.formula1
    assert valor_val.allowBlank is True

    assert parcelas_val is not None
    assert parcelas_val.type == "whole"
    assert parcelas_val.operator == "greaterThanOrEqual"
    assert "1" in parcelas_val.formula1
    assert parcelas_val.allowBlank is True

    # Validate Competência da fatura data validation
    comp_val = None
    for val in ws.data_validations.dataValidation:
        sqref = val.sqref.__str__()
        if "I" in sqref:
            comp_val = val
            break

    assert comp_val is not None
    assert comp_val.type == "date"
    assert comp_val.operator in (None, "between")
    assert comp_val.allowBlank is True
    assert "1" in comp_val.formula1
    assert "2958465" in comp_val.formula2

    req_dark = GenerationRequest(template_id="finance_personal", year=2026, theme="dark")
    path_dark = generate(req_dark, output_dir)
    assert os.path.exists(path_dark)
    
    # Phase 5C: Dark structural regression check
    wb_dark = openpyxl.load_workbook(path_dark, data_only=False)
    assert "Faturas" in wb_dark.sheetnames
    
    ws_faturas_dark = wb_dark["Faturas"]
    assert "tblFaturas" in ws_faturas_dark.tables
    
    table_f_dark = ws_faturas_dark.tables["tblFaturas"]
    assert table_f_dark.ref == "B4:M5" 

    # Phase 5D: Dark structural regression check for Parcelamentos
    assert "Parcelamentos" in wb_dark.sheetnames
    ws_parcelamentos_dark = wb_dark["Parcelamentos"]
    assert "tblParcelamentos" in ws_parcelamentos_dark.tables
    assert ws_parcelamentos_dark.tables["tblParcelamentos"].ref == "B4:O104"

    # Phase 6: Dark structural regression check for Orçamento
    assert "Orçamento" in wb_dark.sheetnames
    ws_orcamento_dark = wb_dark["Orçamento"]
    assert "tblOrcamento" in ws_orcamento_dark.tables
    assert ws_orcamento_dark.tables["tblOrcamento"].ref == "B4:L5"

    # Phase 7: Dark structural regression check for Metas
    assert "Metas" in wb_dark.sheetnames
    ws_metas_dark = wb_dark["Metas"]
    assert "tblMetas" in ws_metas_dark.tables
    assert ws_metas_dark.tables["tblMetas"].ref == "B4:K5"

    # Phase 8: Dark structural regression check for Reserva
    assert "Reserva" in wb_dark.sheetnames
    ws_reserva_dark = wb_dark["Reserva"]
    assert "tblReserva" in ws_reserva_dark.tables
    assert ws_reserva_dark.tables["tblReserva"].ref == "B4:J5"

    # Phase 9: Dark structural regression check for Investimentos
    assert "Investimentos" in wb_dark.sheetnames
    ws_invest_dark = wb_dark["Investimentos"]
    assert "tblInvestimentos" in ws_invest_dark.tables
    assert ws_invest_dark.tables["tblInvestimentos"].ref == "B4:O5"

    # Phase 10: Dark structural regression check for Dashboard Investimentos
    assert "Dashboard Investimentos" in wb_dark.sheetnames
    ws_dash_inv_dark = wb_dark["Dashboard Investimentos"]
    assert "tblResumoInvestimentos" in ws_dash_inv_dark.tables
    assert ws_dash_inv_dark.tables["tblResumoInvestimentos"].ref == "B11:H19"
    assert ws_dash_inv_dark.protection.sheet is True
    # Phase 11: Dark structural regression check for Patrimônio
    assert "Patrimônio" in wb_dark.sheetnames
    ws_patrimonio_dark = wb_dark["Patrimônio"]
    assert "tblBensPatrimoniais" in ws_patrimonio_dark.tables
    assert ws_patrimonio_dark.tables["tblBensPatrimoniais"].ref == "B4:G5"
    
    covered_hidden_pat_dark = set()
    for _, col_dim in ws_patrimonio_dark.column_dimensions.items():
        if getattr(col_dim, "hidden", False):
            for idx in range(col_dim.min, col_dim.max + 1):
                covered_hidden_pat_dark.add(idx)
    assert column_index_from_string("G") in covered_hidden_pat_dark

    # Phase 12: Dark structural regression check for Dashboard Patrimônio
    assert "Dashboard Patrimônio" in wb_dark.sheetnames
    ws_dash_pat_dark = wb_dark["Dashboard Patrimônio"]
    assert "tblResumoPatrimonio" in ws_dash_pat_dark.tables
    assert ws_dash_pat_dark.tables["tblResumoPatrimonio"].ref == "B11:D14"
    assert ws_dash_pat_dark.protection.sheet is True
    
    # Phase 13: Dark structural regression check for Dívidas
    assert "Dívidas" in wb_dark.sheetnames
    ws_dividas_dark = wb_dark["Dívidas"]
    assert "tblDividas" in ws_dividas_dark.tables
    assert ws_dividas_dark.tables["tblDividas"].ref == "B4:K5"
    
    covered_hidden_div_dark = set()
    for _, col_dim in ws_dividas_dark.column_dimensions.items():
        if getattr(col_dim, "hidden", False):
            for idx in range(col_dim.min, col_dim.max + 1):
                covered_hidden_div_dark.add(idx)
    assert column_index_from_string("J") in covered_hidden_div_dark
    assert column_index_from_string("K") in covered_hidden_div_dark

def test_literal_string_starts_with_equal(tmp_path):
    from excel_saas.core.models.workbook_plan import WorkbookPlan, WorksheetPlan, CellPlan
    from excel_saas.core.engine.workbook_engine import WorkbookEngine
    from excel_saas.themes.light import LightTheme

    plan = WorkbookPlan(
        worksheets=[
            WorksheetPlan(
                name="Test",
                cells=[
                    CellPlan(row=0, col=0, value="=not_a_formula"),
                ]
            )
        ]
    )

    engine = WorkbookEngine(plan, LightTheme())
    path = os.path.join(str(tmp_path), "test_string.xlsx")
    engine.render(path)

    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["Test"]

    cell = ws["A1"]
    assert cell.value == "=not_a_formula"
    assert cell.data_type == "s" # String type, not formula ('f')

def test_empty_mode_tables(tmp_path):
    output_dir = str(tmp_path)
    req = GenerationRequest(template_id="finance_personal", year=2026, with_sample_data=False)
    path = generate(req, output_dir)
    wb = openpyxl.load_workbook(path, data_only=False)

    # 1 row of headers, 1 row of blank data
    assert wb["Contas"].tables["tblContas"].ref == "B4:I5"
    assert wb["Cartões"].tables["tblCartoes"].ref == "B6:J7"

    # Assert data validations exist for the blank row in empty mode
    contas_ws = wb["Contas"]
    cartoes_ws = wb["Cartões"]
    # Verify validation is expanded
    assert len(contas_ws.data_validations.dataValidation) > 0
    assert len(cartoes_ws.data_validations.dataValidation) > 0

def test_sample_mode_tables(tmp_path):
    output_dir = str(tmp_path)
    req = GenerationRequest(template_id="finance_personal", year=2026, with_sample_data=True)
    path = generate(req, output_dir)
    wb = openpyxl.load_workbook(path, data_only=False)

    contas_ws = wb["Contas"]
    contas_ref = contas_ws.tables["tblContas"].ref
    # B4:I7 since default sample has 3 accounts (header + 3 data rows)
    assert contas_ref == "B4:I7"

    cartoes_ws = wb["Cartões"]
    cartoes_ref = cartoes_ws.tables["tblCartoes"].ref
    # B6:I8 since default sample has 2 cards (header + 2 data rows)
    assert cartoes_ref == "B6:J8"

    # Sample mode verification: check if Nubank is present
    assert contas_ws["B5"].value == "Nubank"
    assert contas_ws["C5"].value == "Conta corrente"
    assert contas_ws["D5"].value == "Nubank"
    assert contas_ws["E5"].value == 5200.0

    saldo_atual_cell = contas_ws["F5"]
    assert str(saldo_atual_cell.value).startswith("=")
    assert saldo_atual_cell.data_type == "f"

    assert contas_ws["G5"].value == "Sim"
    assert contas_ws["H5"].value == "Sim"

    status_cell = contas_ws["I5"]
    assert str(status_cell.value).startswith("=")
    assert status_cell.data_type == "f"

    # Verify Reserva row
    assert contas_ws["B6"].value == "Reserva"
    assert contas_ws["C6"].value == "Poupança"
    assert contas_ws["D6"].value == "Itaú"
    assert contas_ws["E6"].value == 18000.0
    assert str(contas_ws["F6"].value).startswith("=")
    assert contas_ws["F6"].data_type == "f"
    assert contas_ws["G6"].value == "Não"
    assert contas_ws["H6"].value == "Sim"
    assert str(contas_ws["I6"].value).startswith("=")
    assert contas_ws["I6"].data_type == "f"

    # Verify Nubank card
    assert cartoes_ws["B7"].value == "Nubank"
    assert cartoes_ws["C7"].value == 8000.0
    assert cartoes_ws["D7"].value == 25
    assert cartoes_ws["E7"].value == 2
    assert cartoes_ws["F7"].value == "Nubank"
    assert cartoes_ws["G7"].value == "Sim"
    assert str(cartoes_ws["H7"].value).startswith("=")
    assert cartoes_ws["H7"].data_type == "f"
    assert str(cartoes_ws["I7"].value).startswith("=")
    assert cartoes_ws["I7"].data_type == "f"
    assert str(cartoes_ws["J7"].value).startswith("=")
    assert cartoes_ws["J7"].data_type == "f"

    # Verify Inter Black card
    assert cartoes_ws["B8"].value == "Inter Black"
    assert cartoes_ws["C8"].value == 12000.0
    assert cartoes_ws["D8"].value == 20
    assert cartoes_ws["E8"].value == 5
    assert cartoes_ws["F8"].value == "Nubank"
    assert cartoes_ws["G8"].value == "Sim"
    assert str(cartoes_ws["H8"].value).startswith("=")
    assert cartoes_ws["H8"].data_type == "f"
    assert str(cartoes_ws["I8"].value).startswith("=")
    assert cartoes_ws["I8"].data_type == "f"
    assert str(cartoes_ws["J8"].value).startswith("=")
    assert cartoes_ws["J8"].data_type == "f"

    from openpyxl.utils import column_index_from_string
    col_j_idx = column_index_from_string("J")
    j_hidden = False
    for col_dim in cartoes_ws.column_dimensions.values():
        if getattr(col_dim, "hidden", False):
            d_min = getattr(col_dim, "min", None)
            d_max = getattr(col_dim, "max", None)
            if d_min and d_max and d_min <= col_j_idx <= d_max:
                j_hidden = True
                break
    assert j_hidden, "Column J in Cartões is not physically hidden"
