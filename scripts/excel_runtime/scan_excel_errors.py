import sys
import openpyxl
from openpyxl.utils import range_boundaries

KNOWN_ERRORS = {"#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A"}

def is_cell_in_table_data_row(cell_f, tables):
    """Returns (is_in_table, table_min_col, table_max_col) for table data rows only."""
    for tbl in tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
        if min_row <= cell_f.row <= max_row and min_col <= cell_f.column <= max_col:
            header_rows = tbl.headerRowCount if tbl.headerRowCount is not None else 1
            totals_rows = tbl.totalsRowCount if tbl.totalsRowCount is not None else 0
            if (min_row + header_rows) <= cell_f.row <= (max_row - totals_rows):
                return True, min_col, max_col
    return False, None, None

def scan_workbook(filepath):
    print(f"Scanning {filepath} for Excel formula errors...")

    # Load twice to map formulas to cached values
    wb_formulas = openpyxl.load_workbook(filepath, data_only=False)
    wb_values = openpyxl.load_workbook(filepath, data_only=True)

    errors_found = []
    formulas_scanned = 0

    for sheet_name in wb_formulas.sheetnames:
        ws_f = wb_formulas[sheet_name]
        ws_v = wb_values[sheet_name]

        for row in ws_f.iter_rows():
            for cell_f in row:
                if cell_f.data_type == 'f' or (isinstance(cell_f.value, str) and cell_f.value.startswith('=')):
                    formulas_scanned += 1
                    cell_v = ws_v[cell_f.coordinate]

                    if isinstance(cell_v.value, str) and cell_v.value in KNOWN_ERRORS:
                        ignore_error = False
                        
                        is_table, min_col, max_col = is_cell_in_table_data_row(cell_f, ws_f.tables)
                        if is_table:
                            is_empty_table_row = True
                            for c in range(min_col, max_col + 1):
                                tc = ws_f.cell(row=cell_f.row, column=c)
                                if tc.data_type != 'f' and not (isinstance(tc.value, str) and tc.value.startswith('=')):
                                    if tc.value not in (None, ""):
                                        is_empty_table_row = False
                                        break
                            if is_empty_table_row:
                                ignore_error = True

                        if not ignore_error:
                            errors_found.append({
                                "sheet": sheet_name,
                                "cell": cell_f.coordinate,
                                "formula": cell_f.value,
                                "result": cell_v.value
                            })

    if errors_found:
        print(f"\nFOUND {len(errors_found)} ERRORS in {filepath}:")
        for err in errors_found:
            print(f"{filepath} | {err['sheet']}!{err['cell']}")
            print(f"formula: {err['formula']}")
            print(f"result: {err['result']}\n")
        return False, formulas_scanned, len(errors_found)

    print(f"Scan complete. {formulas_scanned} formulas checked. 0 errors.")
    return True, formulas_scanned, 0

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: scan_excel_errors.py <workbook.xlsx>")
        sys.exit(1)

    filepath = sys.argv[1]
    success, scanned, errors = scan_workbook(filepath)
    if not success:
        sys.exit(1)
    sys.exit(0)
